##############################################################################
#
# Simple Python program to benchmark several Python Excel writing modules.
#
# python bench_excel_writers.py [num_rows] [num_cols]
#
# Copyright 2013-2018, John McNamara, Charlie Clark
#
import functools
import sys
from random import randint, sample
from string import ascii_letters
from time import process_time

import numpy as np
import openpyxl
import pandas as pd
import polars as pl
import xlsxwriter

# Default to 1 sheet with 1000 rows x 50 cols
# 10 % strings

ROW_MAX = 100_000
COL_MAX = 20


def random_string():
    return "".join(sample(ascii_letters, randint(3, 12)))


def _column(i):
    if i < 5:
        return pd.Series([random_string() for _ in range(ROW_MAX)], dtype="string[pyarrow]")
    elif i < 10:
        return np.around(np.random.rand(ROW_MAX) * 10, decimals=2)
    elif i < 15:
        return np.around(np.random.rand(ROW_MAX) * 100, decimals=2)
    return np.random.randint(2, size=ROW_MAX)


df = pd.DataFrame({f"row{i}": np.random.permutation(_column(i)) for i in range(COL_MAX)})

dl = pl.from_dataframe(df)


# df.to_excel("test.xlsx", index=False)


def print_elapsed_time(module_name, elapsed, optimised=False):
    """Print module run times in a consistent format."""
    if optimised:
        module_name += "(optimised)"
    print("    %-28s: %6.2fs" % (module_name.replace("time_", ""), elapsed))


def timer(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = process_time()
        res = func(*args, **kwargs)
        elapsed = process_time() - start_time
        print_elapsed_time(func.__name__, elapsed, kwargs.get("optimised", False))
        return res

    return wrapper


@timer
def pandas_xlsxwriter():
    df.to_excel("pandas_xlsxwriter.xlsx", index=False, engine="xlsxwriter")


@timer
def polars_xlsxwriter():
    dl.write_excel("polars_xlsxwriter.xlsx")


@timer
def pandas_openpyxl():
    df.to_excel("pandas_openpyxl.xlsx", index=False, engine="openpyxl")


@timer
def native_xlsxwriter(optimised=False):
    """Run XlsxWriter in optimised/constant memory mode."""
    filename = f"native_xlsxwriter{int(optimised)}.xlsx"
    workbook = xlsxwriter.Workbook(filename, options=dict(constant_memory=optimised))
    worksheet = workbook.add_worksheet()

    worksheet.write_row(0, 0, list(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=1):
        worksheet.write_row(i, 0, row)

    workbook.close()


@timer
def native_openpyxl(optimised=False):
    """Run OpenPyXL in default mode."""
    filename = f"native_openpyxl{int(optimised)}.xlsx"
    workbook = openpyxl.Workbook(write_only=optimised)
    worksheet = workbook.create_sheet()

    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False):
        worksheet.append(row)

    workbook.save(filename)


print("")
print("Versions:")
print("%s: %s" % ("python", sys.version))
print("%s: %s" % ("openpyxl", openpyxl.__version__))
print("%s: %s" % ("xlsxwriter", xlsxwriter.__version__))
print("")

print("Dimensions:")
print("    Rows = %d" % ROW_MAX)
print("    Cols = %d" % COL_MAX)
print("")

print("Times:")
pandas_openpyxl()
pandas_xlsxwriter()
polars_xlsxwriter()
native_xlsxwriter()
native_xlsxwriter(optimised=True)
native_openpyxl()
native_openpyxl(optimised=True)
print("")
